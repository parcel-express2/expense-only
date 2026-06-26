from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os, json

app = Flask(__name__)
app.config['SECRET_KEY'] = 'expense-only-secret-2024'

# PostgreSQL on Railway, SQLite locally
db_url = os.environ.get('DATABASE_URL', '')
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url or 'sqlite:///expenses.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

CATEGORIES = [
    ('housing',       'السكن',               '🏠'),
    ('food',          'المطاعم',             '🍔'),
    ('groceries',     'البقالة',             '🛒'),
    ('coffee',        'القهوة',              '☕'),
    ('petrol',        'البترول',             '⛽'),
    ('carwash',       'غسيل السيارة',        '🚿'),
    ('carmaint',      'صيانة السيارة',       '🔧'),
    ('health',        'الصحة',               '💊'),
    ('pharmacy',      'الصيدلية',            '💉'),
    ('education',     'التعليم',             '📚'),
    ('entertainment', 'الترفيه',             '🎬'),
    ('clothing',      'الملابس',             '👔'),
    ('internet',      'الهاتف',              '📱'),
    ('subscriptions', 'الاشتراكات',          '📺'),
    ('gifts',         'الهدايا',             '🎁'),
    ('travel',        'السفر',               '✈️'),
    ('personal',      'العناية الشخصية',     '🧴'),
    ('other',         'أخرى',               '📦'),
]


class User(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100), nullable=False, unique=True)
    pin_hash   = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expenses   = db.relationship('Expense', backref='user', lazy=True, cascade='all, delete-orphan')
    budgets    = db.relationship('Budget', backref='user', lazy=True, cascade='all, delete-orphan')


class Expense(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    amount      = db.Column(db.Float, nullable=False)
    category    = db.Column(db.String(50), nullable=False)
    description = db.Column(db.String(200))
    date        = db.Column(db.Date, default=date.today)
    month       = db.Column(db.Integer)
    year        = db.Column(db.Integer)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        if self.date:
            self.month = self.date.month
            self.year  = self.date.year


class PushSub(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    endpoint   = db.Column(db.Text, nullable=False, unique=True)
    sub_json   = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Budget(db.Model):
    id      = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    month   = db.Column(db.Integer, nullable=False)
    year    = db.Column(db.Integer, nullable=False)
    amount  = db.Column(db.Float, nullable=False)


with app.app_context():
    db.create_all()


def get_cat(key):
    for c in CATEGORIES:
        if c[0] == key:
            return c
    return (key, key, '📦')


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ── AUTH ──────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        name = request.form['name'].strip()
        pin  = request.form['pin'].strip()
        user = User.query.filter_by(name=name).first()
        if user and check_password_hash(user.pin_hash, pin):
            session['user_id']   = user.id
            session['user_name'] = user.name
            return redirect(url_for('index'))
        flash('الاسم أو الرمز غير صحيح ❌', 'error')
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name'].strip()
        pin  = request.form['pin'].strip()
        if len(pin) != 4 or not pin.isdigit():
            flash('الرمز يجب أن يكون 4 أرقام ❌', 'error')
            return render_template('register.html')
        if User.query.filter_by(name=name).first():
            flash('هذا الاسم مستخدم بالفعل ❌', 'error')
            return render_template('register.html')
        user = User(name=name, pin_hash=generate_password_hash(pin))
        db.session.add(user)
        db.session.commit()
        session['user_id']   = user.id
        session['user_name'] = user.name
        flash(f'مرحباً {name}! تم إنشاء حسابك 🎉', 'success')
        return redirect(url_for('index'))
    return render_template('register.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── MAIN ──────────────────────────────────────────
@app.route('/')
@login_required
def index():
    user_id = session['user_id']
    today   = date.today()
    month   = request.args.get('month', today.month, type=int)
    year    = request.args.get('year',  today.year,  type=int)
    search  = request.args.get('q', '').strip()

    query = Expense.query.filter_by(user_id=user_id, month=month, year=year)
    if search:
        query = query.filter(
            db.or_(
                Expense.description.ilike(f'%{search}%'),
                Expense.category.ilike(f'%{search}%')
            )
        )
    expenses = query.all()
    total    = sum(e.amount for e in expenses)

    # all expenses this month (without search filter) for stats
    all_expenses = Expense.query.filter_by(user_id=user_id, month=month, year=year).all()
    all_total    = sum(e.amount for e in all_expenses)

    by_cat = {}
    for e in all_expenses:
        by_cat[e.category] = by_cat.get(e.category, 0) + e.amount

    # budget
    budget = Budget.query.filter_by(user_id=user_id, month=month, year=year).first()
    budget_amount = budget.amount if budget else None
    budget_pct    = round(all_total / budget_amount * 100, 1) if budget_amount else None

    # prev month comparison
    prev_month = month - 1 if month > 1 else 12
    prev_year  = year if month > 1 else year - 1
    prev_total = db.session.query(db.func.sum(Expense.amount))\
                   .filter_by(user_id=user_id, month=prev_month, year=prev_year).scalar() or 0
    diff_pct = round((all_total - prev_total) / prev_total * 100, 1) if prev_total > 0 else None

    arabic_months = {
        1:'يناير',2:'فبراير',3:'مارس',4:'أبريل',
        5:'مايو',6:'يونيو',7:'يوليو',8:'أغسطس',
        9:'سبتمبر',10:'أكتوبر',11:'نوفمبر',12:'ديسمبر'
    }
    years_list = list(range(today.year - 2, today.year + 2))

    return render_template('index.html',
        expenses=expenses,
        total=total,
        all_total=all_total,
        by_cat=by_cat,
        categories=CATEGORIES,
        current_month=month,
        current_year=year,
        years_list=years_list,
        arabic_months=arabic_months,
        get_cat=get_cat,
        user_name=session.get('user_name'),
        search=search,
        budget_amount=budget_amount,
        budget_pct=budget_pct,
        diff_pct=diff_pct,
        prev_month=prev_month,
        prev_total=prev_total,
    )


@app.route('/add_expense', methods=['POST'])
@login_required
def add_expense():
    amount      = float(request.form['amount'])
    category    = request.form['category']
    description = request.form.get('description', '')
    entry_date  = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
    e = Expense(user_id=session['user_id'], amount=amount,
                category=category, description=description, date=entry_date)
    db.session.add(e)
    db.session.commit()
    flash('تمت إضافة المصروف ✅', 'success')
    return redirect(url_for('index', month=entry_date.month, year=entry_date.year))


@app.route('/edit_expense/<int:id>', methods=['POST'])
@login_required
def edit_expense(id):
    e = Expense.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    e.amount      = float(request.form['amount'])
    e.category    = request.form['category']
    e.description = request.form.get('description', '')
    entry_date    = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
    e.date  = entry_date
    e.month = entry_date.month
    e.year  = entry_date.year
    db.session.commit()
    flash('تم التعديل ✅', 'success')
    return redirect(url_for('index', month=e.month, year=e.year))


@app.route('/delete_expense/<int:id>')
@login_required
def delete_expense(id):
    e = Expense.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    month, year = e.month, e.year
    db.session.delete(e)
    db.session.commit()
    flash('تم الحذف ✅', 'info')
    return redirect(url_for('index', month=month, year=year))


@app.route('/set_budget', methods=['POST'])
@login_required
def set_budget():
    month  = int(request.form['month'])
    year   = int(request.form['year'])
    amount = float(request.form['budget_amount'])
    budget = Budget.query.filter_by(user_id=session['user_id'], month=month, year=year).first()
    if budget:
        budget.amount = amount
    else:
        budget = Budget(user_id=session['user_id'], month=month, year=year, amount=amount)
        db.session.add(budget)
    db.session.commit()
    flash('تم حفظ الميزانية ✅', 'success')
    return redirect(url_for('index', month=month, year=year))


@app.route('/export_excel')
@login_required
def export_excel():
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('مكتبة التصدير غير متاحة', 'error')
        return redirect(url_for('index'))

    user_id = session['user_id']
    today   = date.today()
    month   = request.args.get('month', today.month, type=int)
    year    = request.args.get('year',  today.year,  type=int)

    expenses = Expense.query.filter_by(user_id=user_id, month=month, year=year)\
                            .order_by(Expense.date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    arabic_months = {1:'يناير',2:'فبراير',3:'مارس',4:'أبريل',5:'مايو',6:'يونيو',
                     7:'يوليو',8:'أغسطس',9:'سبتمبر',10:'أكتوبر',11:'نوفمبر',12:'ديسمبر'}
    ws.title = f"{arabic_months[month]} {year}"

    # Header
    headers = ['التاريخ', 'الفئة', 'الوصف', 'المبلغ (ر.ع)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='C0392B')
        cell.alignment = Alignment(horizontal='center')

    for row, e in enumerate(expenses, 2):
        cat = get_cat(e.category)
        ws.cell(row=row, column=1, value=str(e.date))
        ws.cell(row=row, column=2, value=f"{cat[2]} {cat[1]}")
        ws.cell(row=row, column=3, value=e.description or '')
        ws.cell(row=row, column=4, value=round(e.amount, 3))

    # Total row
    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=3, value='الإجمالي').font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=4, value=round(sum(e.amount for e in expenses), 3))
    total_cell.font = Font(bold=True, color='C0392B')

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"مصروفات_{arabic_months[month]}_{year}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/chart')
@login_required
def chart():
    user_id = session['user_id']
    today   = date.today()
    mode    = request.args.get('mode', 'monthly')
    year    = request.args.get('year',  today.year,  type=int)
    month   = request.args.get('month', today.month, type=int)

    if mode == 'daily':
        import calendar
        days_in_month = calendar.monthrange(year, month)[1]
        labels = [str(d) for d in range(1, days_in_month + 1)]
        data = []
        for d in range(1, days_in_month + 1):
            try:
                day_date = date(year, month, d)
            except ValueError:
                data.append(0)
                continue
            total = db.session.query(db.func.sum(Expense.amount))\
                      .filter(Expense.user_id == user_id, Expense.date == day_date).scalar() or 0
            data.append(round(total, 3))
    else:
        labels = ['يناير','فبراير','مارس','أبريل','مايو','يونيو',
                  'يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']
        data = []
        for m in range(1, 13):
            total = db.session.query(db.func.sum(Expense.amount))\
                      .filter_by(user_id=user_id, month=m, year=year).scalar() or 0
            data.append(round(total, 3))

    return jsonify({'labels': labels, 'data': data, 'mode': mode})


@app.route('/scan_receipt', methods=['POST'])
@login_required
def scan_receipt():
    import requests as http_requests
    from PIL import Image
    import io, json, re, base64

    api_key = os.environ.get('MISTRAL_API_KEY') or os.environ.get('GEMINI_API_KEY')
    if not api_key:
        return jsonify({'error': 'مفتاح AI غير موجود'}), 500

    if 'image' not in request.files:
        return jsonify({'error': 'لم يتم إرسال صورة'}), 400

    file = request.files['image']
    img_bytes = file.read()

    try:
        try:
            import pillow_heif
            pillow_heif.register_heif_opener()
        except ImportError:
            pass

        img = Image.open(io.BytesIO(img_bytes))
        if img.mode != 'RGB':
            img = img.convert('RGB')

        max_size = 1600
        if max(img.size) > max_size:
            ratio = max_size / max(img.size)
            img = img.resize((int(img.size[0]*ratio), int(img.size[1]*ratio)), Image.LANCZOS)

        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=85)
        jpeg_b64 = base64.b64encode(buf.getvalue()).decode('utf-8')

        prompt = (
            "You are a receipt reader. Look at this receipt image carefully.\n"
            "Reply with ONLY a raw JSON object, no markdown, no explanation:\n"
            '{"amount": <total amount as number>, '
            '"description": "<store name or description in Arabic>", '
            '"category": "<one of: food, groceries, coffee, petrol, carwash, carmaint, health, pharmacy, education, entertainment, clothing, utilities, internet, subscriptions, savings, gifts, travel, housing, other>", '
            '"date": "<date as YYYY-MM-DD or empty string>"}\n\n'
            "Category rules: coffee shop/cafe→coffee, gas station/fuel→petrol, restaurant→food, "
            "supermarket/grocery→groceries, pharmacy→pharmacy, hospital/clinic→health, "
            "clothes/shoes→clothing, electricity/water bill→utilities, carwash→carwash, "
            "car repair/parts→carmaint, otherwise→other. "
            "If amount unclear use 0. Write description in Arabic."
        )

        url = "https://api.mistral.ai/v1/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        mistral_payload = {
            "model": "pixtral-12b-2409",
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": f"data:image/jpeg;base64,{jpeg_b64}"}
                ]
            }],
            "max_tokens": 300
        }
        resp = http_requests.post(url, json=mistral_payload, headers=headers, timeout=30)
        resp.raise_for_status()
        result = resp.json()
        text = result['choices'][0]['message']['content'].strip()

        text = re.sub(r'```(?:json)?', '', text).strip()
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            data = json.loads(match.group())
        else:
            data = json.loads(text)

        try:
            data['amount'] = float(str(data.get('amount', 0)).replace(',', '.'))
        except Exception:
            data['amount'] = 0

        return jsonify({'success': True, 'data': data})

    except Exception as e:
        return jsonify({'error': str(e), 'detail': 'scan_failed'}), 500




import re as _re

@app.route('/sms_webhook', methods=['POST'])
def sms_webhook():
    """
    iOS Shortcuts يرسل رسالة SMS من bankmuscat هنا تلقائياً.
    JSON body: { "message": "...", "username": "...", "pin": "..." }
    """
    data = request.get_json(silent=True) or {}
    message_body = data.get('message', '')
    username     = data.get('username', '')
    pin          = data.get('pin', '')

    if not message_body:
        return jsonify({'error': 'no message'}), 400

    # فلتر — فقط رسائل الخصم
    if 'تم خصم' not in message_body:
        return jsonify({'ignored': 'not a debit message'}), 200

    # التحقق من المستخدم
    user = User.query.filter_by(name=username).first()
    if not user or not check_password_hash(user.pin_hash, str(pin)):
        return jsonify({'error': 'unauthorized'}), 401

    # استخراج المبلغ من رسالة بنك مسقط
    # صيغة بنك مسقط: "تم خصم 0.400 OMR" أو "OMR 12.500" أو "RO 5.000"
    amount = None
    amt_match = _re.search(r'([\d,]+\.\d+)\s*(?:OMR|RO)|(?:OMR|RO)\s*([\d,]+\.\d+)', message_body, _re.IGNORECASE)
    if amt_match:
        try:
            raw = amt_match.group(1) or amt_match.group(2)
            amount = float(raw.replace(',', ''))
        except Exception:
            amount = None

    if not amount or amount <= 0:
        return jsonify({'error': 'could not parse amount', 'message': message_body}), 422

    # تصنيف تلقائي
    lower = message_body.lower()
    if any(w in lower for w in ['coffee', 'cafe', 'tea', 'juice', 'قهوة', 'شاي', 'عصير']):
        cat = 'coffee'
    elif any(w in lower for w in ['restaurant', 'مطعم', 'burger', 'pizza', 'shawarma', 'grill']):
        cat = 'food'
    elif any(w in lower for w in ['petrol', 'fuel', 'station', 'بترول', 'وقود', 'محطة']):
        cat = 'petrol'
    elif any(w in lower for w in ['pharmacy', 'medical', 'hospital', 'clinic', 'صيدلية', 'طبي', 'مستشفى']):
        cat = 'pharmacy'
    elif any(w in lower for w in ['supermarket', 'lulu', 'carrefour', 'hypermarket', 'market', 'بقالة', 'سوبرماركت']):
        cat = 'groceries'
    else:
        cat = 'other'

    # استخراج اسم المتجر من رسالة بنك مسقط
    # صيغة 1: "في 901279-AL JAZEERA TEA AL K بتاريخ"
    # صيغة 2: "في AL BURG AL BRONZE TRAD AL KHABOURA OM بتاريخ"
    desc_match = _re.search(r'في\s+(?:[\d\w]+-)?(.+?)\s+بتاريخ', message_body)
    if desc_match:
        description = desc_match.group(1).strip()
    else:
        description = 'بنك مسقط - دفعة تلقائية'

    expense = Expense(
        user_id     = user.id,
        amount      = amount,
        category    = cat,
        description = description,
        date        = date.today()
    )
    db.session.add(expense)
    db.session.commit()

    return jsonify({
        'success': True,
        'added': {
            'amount':      amount,
            'category':    cat,
            'description': description
        }
    }), 200


VAPID_PRIVATE = os.environ.get('VAPID_PRIVATE_KEY', '')
VAPID_PUBLIC  = os.environ.get('VAPID_PUBLIC_KEY', '')
VAPID_EMAIL   = 'mailto:admin@masarify.app'


@app.route('/api/vapid_public')
def vapid_public():
    return jsonify({'key': VAPID_PUBLIC})


@app.route('/api/subscribe_push', methods=['POST'])
@login_required
def subscribe_push():
    sub = request.get_json()
    if not sub or 'endpoint' not in sub:
        return jsonify({'error': 'invalid'}), 400
    existing = PushSub.query.filter_by(endpoint=sub['endpoint']).first()
    if existing:
        existing.sub_json = json.dumps(sub)
        existing.user_id  = session['user_id']
    else:
        db.session.add(PushSub(
            user_id  = session['user_id'],
            endpoint = sub['endpoint'],
            sub_json = json.dumps(sub)
        ))
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/unsubscribe_push', methods=['POST'])
@login_required
def unsubscribe_push():
    data = request.get_json() or {}
    PushSub.query.filter_by(user_id=session['user_id'], endpoint=data.get('endpoint','')).delete()
    db.session.commit()
    return jsonify({'ok': True})


def send_push(user_id, title, body):
    """إرسال إشعار لجميع أجهزة المستخدم"""
    if not VAPID_PRIVATE or not VAPID_PUBLIC:
        return
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        return
    subs = PushSub.query.filter_by(user_id=user_id).all()
    for s in subs:
        try:
            webpush(
                subscription_info   = json.loads(s.sub_json),
                data                = json.dumps({'title': title, 'body': body}),
                vapid_private_key   = VAPID_PRIVATE,
                vapid_claims        = {'sub': VAPID_EMAIL}
            )
        except Exception:
            db.session.delete(s)
    db.session.commit()


@app.route('/api/check_budget_notify')
@login_required
def check_budget_notify():
    """يُستدعى من الـ frontend عند فتح التطبيق للتحقق من الميزانية"""
    user_id = session['user_id']
    today   = date.today()
    budget  = Budget.query.filter_by(user_id=user_id, month=today.month, year=today.year).first()
    if not budget:
        return jsonify({'ok': True, 'no_budget': True})
    total = db.session.query(db.func.sum(Expense.amount))\
              .filter_by(user_id=user_id, month=today.month, year=today.year).scalar() or 0
    pct = total / budget.amount * 100
    if pct >= 100:
        send_push(user_id, '⚠️ تجاوزت الميزانية!',
                  f'صرفت {total:.3f} ر.ع من {budget.amount:.3f} ر.ع')
        return jsonify({'alert': 'exceeded', 'pct': round(pct,1)})
    elif pct >= 90:
        send_push(user_id, '🔔 اقتربت من الميزانية',
                  f'استهلكت {pct:.0f}% من ميزانية الشهر')
        return jsonify({'alert': 'warning', 'pct': round(pct,1)})
    return jsonify({'ok': True, 'pct': round(pct,1)})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5051))
    app.run(host='0.0.0.0', debug=True, port=port)
